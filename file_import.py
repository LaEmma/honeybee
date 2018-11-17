from collections import OrderedDict
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import save_virtual_workbook


class ExcelImporter(object):
    """
    认为导入文件每行应该是一个对象
    最后返回以key为excel中行数的字典

    imported_fields如果提供，则可以根据提供的field返回
    """
    def __init__(self, imported_fields={}):
        super(ExcelImporter, self).__init__()
        self.imported_fields = imported_fields

    def import_file(self, file):
        '''
        只解析文件，解析后返回一个字典，每一行应该是一个对象
        '''
        result = OrderedDict()
        # 记录所在列 fields_columns = {'title':'D', 'name':'E', 'desc':'F'}
        fields_columns = dict()
        wb = load_workbook(file, read_only=False, data_only=True)
        # 模板里只用一页
        ws = wb.worksheets[0]
        # 把每列对应什么字段记录在fields_columns
        for column in ws.iter_cols():
            for field_key, field_value in self.imported_fields.items():
                # 认为只有第一行是title
                if column[0].value == field_value.strip():
                    fields_columns.update({
                        field_key: column[0].column # 'title': 'D'
                    })
                    break
        # 按照每行读入
        # 如果有指定的fields
        if self.imported_fields:
            for row_index in range(2, ws.max_row + 1):
                fields_row = OrderedDict()
                for field_name, column_index in fields_columns.items():
                    value = ws['%s%s' % (column_index, row_index)].value
                    if value is not None:
                        field = {field_name: ws['%s%s' % (column_index, row_index)].value}
                        fields_row.update(field)
                result.update({
                    str(row_index): fields_row
                })
        # 如果没有指定fields则全返回
        else:
            for row_index in range(ws.min_row, ws.max_row + 1):
                fields_row = OrderedDict()
                for column_index in range(ws.min_column - 1, ws.max_column):
                    cell = ws[row_index][column_index]
                    field = {cell.column: cell.value}
                    fields_row.update(field)
                result.update({
                    str(row_index): fields_row
                })
        return result

    def export_file(self):
        pass


# class ExcelExporter(object):
#     """
#     exported_fields 应该是一个字典，内容是属性或者变量所对应的字段
#     e.g.
#         exported_fields = {
#             'title': '商品名称',
#             'sku_attr': '商品属性'
#         }
#     之后
#     items 应该是 products
#     获取每行数据的时候，会迭代传入的items，用exported_fields里的key作为要获取的变量或属性

#     exported_fields 为空的话，items直接输出

#     2018-07-11
#     试验过，发现传进来 exported_fields 再变OrderedDict，是没办法和items顺序一致的
#     所以做的这么复杂不如直接用户拼好数据了

#     """
#     def __init__(self, items, exported_fields={}):
#         super(ExcelImporter, self).__init__()
#         self.exported_fields = OrderedDict(exported_fields)
#         self.items = items

#     def validate_data(self):
#         '''
#         验证 title 的数量 和 data里面一条的数量对的上
#         '''
#         if self.exported_fields:
#             first_data = self.items[0]
#             if len(first_data) != len(self.exported_fields.keys()):
#                 raise ValueError('传入exported_fields和数据items一行的数量不一致。')

#     def export_file(self):
#         '''
#         title 放置导出excel里第一行想展示的文字，可以为空
#         data 应该是一个list，可迭代，每一个元素就是一行
#         '''
#         self.validate_data()
#         wb = Workbook()
#         ws = wb.active
#         # result = []
#         if self.exported_fields:
#             titles = self.exported_fields.values()
#             import pdb; pdb.set_trace()
#             ws.append(titles)
#         for item in self.items:


def export_file(items):
    '''
    '''
    wb = Workbook()
    ws = wb.active
    for item in items:
        ws.append(item)
    result = save_virtual_workbook(wb)
    return result
    # response = HttpResponse(
    #     result, content_type='application/vnd.ms-excel', status=200)
    # return response
